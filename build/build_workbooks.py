# -*- coding: utf-8 -*-
"""
build_workbooks.py - all the Excel artifacts for the Field Notes Lab (fictional Meridian Pay).

Each workbook ships as a filled example and, where useful, a BLANK version to practise on.
Real NIST structure (CSF 2.0 functions and categories, FIPS 199, SP 800-53 families),
fictional company data.

    python build_workbooks.py

Outputs into the episode and template folders.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FONT = "Calibri"
FOREST = "2E5E4E"; DEEP = "173029"; PAPER = "F0F3EC"; SAGE = "DCE4DC"
AMBER = "A9772E"; RUST = "8C4A34"; MUTED = "5F6A60"; WHITE = "FFFFFF"; ROW = "F5F7F2"
thin = Side(style="thin", color="C6D0C2")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

def outpath(*parts):
    p = os.path.join(ROOT, *parts)
    os.makedirs(os.path.dirname(p), exist_ok=True)
    return p

def title_bar(ws, text, span, sub=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name=FONT, bold=True, size=15, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DEEP)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30
    if sub:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
        s = ws.cell(row=2, column=1, value=sub)
        s.font = Font(name=FONT, italic=True, size=10, color=MUTED)

def header_row(ws, r, headers, fill=FOREST):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(name=FONT, bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORD
    ws.row_dimensions[r].height = 30

def body(cell, bold=False, wrap=True, align="left", size=10, color="000000"):
    cell.font = Font(name=FONT, bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = BORD

def widths(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w

# ============================================================ CSF profile
CSF = [
 ("GOVERN","GV.OC","Organizational Context","We understand our mission, legal duties, and what we must protect.",2,4,"High"),
 ("GOVERN","GV.RM","Risk Management Strategy","We have decided how much risk we accept and how we manage it.",2,4,"High"),
 ("GOVERN","GV.RR","Roles, Responsibilities, Authorities","Everyone knows who owns what in security.",2,3,"Medium"),
 ("GOVERN","GV.PO","Policy","Security policies exist, are approved, and are kept current.",3,4,"Medium"),
 ("GOVERN","GV.OV","Oversight","Leadership reviews the programme and adjusts it.",1,3,"High"),
 ("GOVERN","GV.SC","Cybersecurity Supply Chain Risk Mgmt","We manage risk from providers (AWS, CardaX, VerifyNow).",2,4,"High"),
 ("IDENTIFY","ID.AM","Asset Management","We know our systems, data, and who owns them.",2,4,"High"),
 ("IDENTIFY","ID.RA","Risk Assessment","We find and rate our vulnerabilities and threats.",2,4,"High"),
 ("IDENTIFY","ID.IM","Improvement","We learn from tests, incidents, and reviews and improve.",1,3,"Medium"),
 ("PROTECT","PR.AA","Identity, Authentication, Access Control","Only the right people and systems get access, with MFA.",3,4,"Medium"),
 ("PROTECT","PR.AT","Awareness and Training","Staff know how to work safely and spot problems.",2,3,"Medium"),
 ("PROTECT","PR.DS","Data Security","Data is protected at rest and in transit.",3,4,"Medium"),
 ("PROTECT","PR.PS","Platform Security","Systems are configured securely and patched.",2,4,"High"),
 ("PROTECT","PR.IR","Technology Infrastructure Resilience","Systems are built to keep running and resist attack.",2,3,"Medium"),
 ("DETECT","DE.CM","Continuous Monitoring","We watch systems and networks for problems.",2,4,"High"),
 ("DETECT","DE.AE","Adverse Event Analysis","We analyse alerts to understand what is happening.",1,3,"High"),
 ("RESPOND","RS.MA","Incident Management","We have a plan and run it when incidents happen.",2,4,"High"),
 ("RESPOND","RS.AN","Incident Analysis","We investigate incidents to find scope and cause.",1,3,"Medium"),
 ("RESPOND","RS.CO","Incident Reporting and Communication","We tell the right people, inside and out.",2,3,"Medium"),
 ("RESPOND","RS.MI","Incident Mitigation","We contain and reduce the impact of incidents.",2,3,"Medium"),
 ("RECOVER","RC.RP","Incident Recovery Plan Execution","We restore systems and data after an incident.",2,4,"High"),
 ("RECOVER","RC.CO","Incident Recovery Communication","We keep people informed during recovery.",1,3,"Medium"),
]
FUNC_COLOR = {"GOVERN":DEEP,"IDENTIFY":FOREST,"PROTECT":"3A6B57","DETECT":"4A6B5A","RESPOND":RUST,"RECOVER":AMBER}

def csf_profile(blank=False):
    wb = Workbook(); ws = wb.active; ws.title = "CSF Profile"
    title_bar(ws, "Meridian Pay  -  Cybersecurity Framework 2.0 Profile", 9,
              "Rate each category now (Current) and where you want to be (Target). Gap and the summary update themselves. Scale 1-4.")
    hdr = ["Function","ID","Category","Outcome (plain language)","Current (1-4)","Target (1-4)","Gap","Priority","Notes"]
    header_row(ws, 4, hdr)
    r = 5
    for func, cid, cat, outcome, cur, tgt, prio in CSF:
        vals = [func, cid, cat, outcome,
                ("" if blank else cur), ("" if blank else tgt),
                None, ("" if blank else prio), ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            body(cell, align="center" if c in (2,5,6,7,8) else "left")
            if r % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ROW)
        ws.cell(row=r, column=1).font = Font(name=FONT, bold=True, size=10, color=WHITE)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=FUNC_COLOR[func])
        # gap formula
        ws.cell(row=r, column=7).value = f'=IF(AND(ISNUMBER(E{r}),ISNUMBER(F{r})),F{r}-E{r},"")'
        r += 1
    last = r - 1
    # data validations
    dv14 = DataValidation(type="whole", operator="between", formula1=1, formula2=4, allow_blank=True)
    dv14.error = "Enter a whole number 1 to 4."; dv14.prompt = "1 Initial, 2 Developing, 3 Defined, 4 Managed"
    ws.add_data_validation(dv14); dv14.add(f"E5:F{last}")
    dvp = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(dvp); dvp.add(f"H5:H{last}")
    # conditional format gap
    ws.conditional_formatting.add(f"G5:G{last}",
        CellIsRule(operator="greaterThanOrEqual", formula=["2"], fill=PatternFill("solid", fgColor="F3D6CE")))
    ws.conditional_formatting.add(f"G5:G{last}",
        CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor="F6E6C8")))
    ws.conditional_formatting.add(f"G5:G{last}",
        CellIsRule(operator="lessThanOrEqual", formula=["0"], fill=PatternFill("solid", fgColor="D8E6DC")))
    # summary block
    sr = last + 3
    ws.cell(row=sr, column=1, value="Summary by Function").font = Font(name=FONT, bold=True, size=12, color=DEEP)
    header_row(ws, sr+1, ["Function","Avg Current","Avg Target","Avg Gap","",""][:4])
    funcs = ["GOVERN","IDENTIFY","PROTECT","DETECT","RESPOND","RECOVER"]
    rr = sr + 2
    for f in funcs:
        rows = [i+5 for i,row in enumerate(CSF) if row[0]==f]
        rngE = f"E{rows[0]}:E{rows[-1]}"; rngF = f"F{rows[0]}:F{rows[-1]}"; rngG = f"G{rows[0]}:G{rows[-1]}"
        ws.cell(row=rr, column=1, value=f); body(ws.cell(row=rr,column=1), bold=True)
        for col, rng in ((2,rngE),(3,rngF),(4,rngG)):
            cell = ws.cell(row=rr, column=col, value=f'=IFERROR(ROUND(AVERAGE({rng}),1),"")')
            body(cell, align="center"); cell.number_format = "0.0"
        rr += 1
    widths(ws, {"A":11,"B":9,"C":30,"D":46,"E":11,"F":11,"G":7,"H":10,"I":30})
    ws.freeze_panes = "A5"
    # legend sheet
    lg = wb.create_sheet("How to use")
    lines = [
        ("How to use this profile", True),
        ("", False),
        ("1. For each row, set Current: how well Meridian does this today, 1 to 4.", False),
        ("2. Set Target: where you want to be. Be honest, not aspirational everywhere.", False),
        ("3. Gap fills itself. A gap of 2 or more (shaded red) is where to focus first.", False),
        ("4. Use Priority and Notes to plan the work.", False),
        ("", False),
        ("The 1 to 4 scale (a simple maturity rating, echoing CSF Tiers):", True),
        ("1  Initial     Done ad hoc, if at all. No consistency.", False),
        ("2  Developing  Happening, but not standardised or complete.", False),
        ("3  Defined     Documented, consistent, and followed.", False),
        ("4  Managed     Measured, reviewed, and improved over time.", False),
        ("", False),
        ("This profile assesses the 22 CSF 2.0 Categories across all 6 Functions.", False),
        ("The exact Subcategory wording lives in the official CSF 2.0 (CSWP 29) at csrc.nist.gov.", False),
        ("These plain-language outcomes are a teaching paraphrase, not the normative text.", False),
    ]
    for i,(t,b) in enumerate(lines,1):
        cell = lg.cell(row=i, column=1, value=t)
        cell.font = Font(name=FONT, bold=b, size=13 if b and i==1 else 11, color=DEEP if b else "000000")
    lg.column_dimensions["A"].width = 95
    name = "meridian-csf-profile-BLANK.xlsx" if blank else "meridian-csf-profile.xlsx"
    wb.save(outpath("episode-01-csf", name)); print(name)

# ============================================================ FIPS 199
INFO_TYPES = [
 ("Customer profile data shown in portal","name, contact, masked account details","Moderate","Moderate","Moderate"),
 ("Account balances and history (displayed)","read from Core Banking, shown to the customer","Moderate","Moderate","Moderate"),
 ("Session and login tokens","who is logged in right now","Moderate","Moderate","Moderate"),
 ("Portal content and configuration","the pages and settings themselves","Low","Moderate","Moderate"),
]
def fips199(blank=False):
    wb = Workbook(); ws = wb.active; ws.title = "FIPS 199"
    title_bar(ws, "Meridian Pay  -  FIPS 199 Categorisation: Customer Web Portal (S2)", 6,
              "Rate each information type for Confidentiality, Integrity, Availability. The system takes the highest of each (the high-water mark).")
    header_row(ws, 4, ["Information type","What it is","Confidentiality","Integrity","Availability","Rationale"])
    r = 5
    for name, what, c_, i_, a_ in INFO_TYPES:
        vals = [name, what, ("" if blank else c_), ("" if blank else i_), ("" if blank else a_), ""]
        for c, v in enumerate(vals,1):
            cell = ws.cell(row=r, column=c, value=v); body(cell, align="center" if c in (3,4,5) else "left")
            if r%2==0: cell.fill = PatternFill("solid", fgColor=ROW)
        r += 1
    last = r-1
    dvl = DataValidation(type="list", formula1='"Low,Moderate,High"', allow_blank=True)
    ws.add_data_validation(dvl); dvl.add(f"C5:E{last}")
    # result block
    rr = last+2
    ws.cell(row=rr, column=1, value="System categorisation (high-water mark)").font = Font(name=FONT, bold=True, size=12, color=DEEP)
    labels = [("Confidentiality","C"),("Integrity","D"),("Availability","E")]
    for k,(lab,col) in enumerate(labels):
        ws.cell(row=rr+1+k, column=1, value=lab); body(ws.cell(row=rr+1+k,column=1), bold=True)
        formula = (f'=IF(COUNTIF({col}5:{col}{last},"High"),"High",'
                   f'IF(COUNTIF({col}5:{col}{last},"Moderate"),"Moderate",'
                   f'IF(COUNTIF({col}5:{col}{last},"Low"),"Low","")))')
        cell = ws.cell(row=rr+1+k, column=2, value=formula); body(cell, align="center", bold=True)
    orow = rr+5
    ws.cell(row=orow, column=1, value="OVERALL SYSTEM IMPACT").font = Font(name=FONT, bold=True, size=12, color=WHITE)
    ws.cell(row=orow, column=1).fill = PatternFill("solid", fgColor=RUST)
    ov = (f'=IF(OR(B{rr+1}="High",B{rr+2}="High",B{rr+3}="High"),"HIGH",'
          f'IF(OR(B{rr+1}="Moderate",B{rr+2}="Moderate",B{rr+3}="Moderate"),"MODERATE",'
          f'IF(OR(B{rr+1}="Low",B{rr+2}="Low",B{rr+3}="Low"),"LOW","")))')
    oc = ws.cell(row=orow, column=2, value=ov); oc.font = Font(name=FONT, bold=True, size=12, color=RUST)
    oc.alignment = Alignment(horizontal="center")
    note = ws.cell(row=orow+2, column=1,
        value="Why Moderate and not High: the portal shows masked and limited data and hands sensitive operations to the "
              "Core Banking system. Its own boundary is Moderate. The authoritative account data (High) is categorised "
              "separately with Core Banking (S3). This scoping choice is the whole reason a categorisation is written down: "
              "so the sign-off is about a defined boundary, not a vague feeling. The Moderate result drives the SP 800-53B "
              "Moderate baseline you tailor in the control-selection workbook.")
    note.font = Font(name=FONT, italic=True, size=10, color=MUTED); note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=orow+2, start_column=1, end_row=orow+4, end_column=6)
    ws.row_dimensions[orow+2].height = 60
    widths(ws, {"A":30,"B":26,"C":16,"D":14,"E":14,"F":40})
    ws.freeze_panes = "A5"
    name = "fips199-categorization-BLANK.xlsx" if blank else "fips199-categorization.xlsx"
    wb.save(outpath("episode-03-categorize-select", name)); print(name)

# ============================================================ control selection
CONTROLS = [
 ("AC-2","Account Management","Access Control","Yes","Partly","Platform team","Joiner-mover-leaver runs, quarterly review not yet formal"),
 ("AC-3","Access Enforcement","Access Control","Yes","Yes","Platform team","Role-based access in place"),
 ("AC-6","Least Privilege","Access Control","Yes","Partly","Platform team","Some standing admin to reduce"),
 ("AC-17","Remote Access","Access Control","Yes","Yes","Platform team","All staff remote, VPN and MFA"),
 ("AU-2","Event Logging","Audit and Accountability","Yes","Partly","Platform team","Logs collected, coverage gaps on admin console"),
 ("AU-6","Audit Record Review","Audit and Accountability","Yes","No","Security team","No regular review process yet"),
 ("CA-7","Continuous Monitoring","Assessment, Authorization, Monitoring","Yes","No","CISO","Programme being defined"),
 ("CM-2","Baseline Configuration","Configuration Management","Yes","Partly","Platform team","Infrastructure as code, no formal baseline doc"),
 ("CM-6","Configuration Settings","Configuration Management","Yes","Partly","Platform team","Hardening partly automated"),
 ("CP-9","System Backup","Contingency Planning","Yes","Yes","Platform team","Automated, encrypted"),
 ("CP-10","System Recovery","Contingency Planning","Yes","Partly","Platform team","Restores not tested on a schedule"),
 ("IA-2","Identification and Authentication","Identification and Authentication","Yes","Yes","Platform team","MFA for all users"),
 ("IA-5","Authenticator Management","Identification and Authentication","Yes","Yes","Platform team","Password and MFA rules enforced"),
 ("IR-4","Incident Handling","Incident Response","Yes","Partly","CISO","Plan exists, needs rehearsal"),
 ("IR-6","Incident Reporting","Incident Response","Yes","Partly","CISO","Internal path clear, regulator path being documented"),
 ("RA-5","Vulnerability Monitoring","Risk Assessment","Yes","Partly","Security team","Scanning in place, cadence inconsistent"),
 ("SC-7","Boundary Protection","System and Communications Protection","Yes","Yes","Platform team","Web firewall and segmentation"),
 ("SC-8","Transmission Confidentiality","System and Communications Protection","Yes","Yes","Platform team","TLS everywhere"),
 ("SC-13","Cryptographic Protection","System and Communications Protection","Yes","Yes","Platform team","Encryption at rest and in transit"),
 ("SI-2","Flaw Remediation","System and Information Integrity","Yes","Partly","Platform team","Patching driven, SLA not formal"),
 ("SI-4","System Monitoring","System and Information Integrity","Yes","Partly","Security team","Monitoring exists, tuning needed"),
 ("SI-7","Software and Information Integrity","System and Information Integrity","Yes","Partly","Platform team","Some integrity checks"),
]
def control_selection(blank=False):
    wb = Workbook(); ws = wb.active; ws.title = "Control Selection"
    title_bar(ws, "Meridian Pay  -  Control Selection: Customer Web Portal (Moderate baseline)", 7,
              "Start from the SP 800-53B Moderate baseline (the portal categorised as Moderate). Mark what applies, its status, and who owns it.")
    header_row(ws, 4, ["Control ID","Control name","Family","In baseline","Implemented","Owner","Notes"])
    r = 5
    for cid, cname, fam, base, impl, owner, note in CONTROLS:
        vals = [cid, cname, fam, base, ("" if blank else impl), ("" if blank else owner), ("" if blank else note)]
        for c,v in enumerate(vals,1):
            cell = ws.cell(row=r, column=c, value=v); body(cell, align="center" if c in (4,5) else "left")
            if r%2==0: cell.fill = PatternFill("solid", fgColor=ROW)
        r += 1
    last = r-1
    dvi = DataValidation(type="list", formula1='"Yes,Partly,No"', allow_blank=True)
    ws.add_data_validation(dvi); dvi.add(f"E5:E{last}")
    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="equal", formula=['"Yes"'], fill=PatternFill("solid", fgColor="D8E6DC")))
    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="equal", formula=['"Partly"'], fill=PatternFill("solid", fgColor="F6E6C8")))
    ws.conditional_formatting.add(f"E5:E{last}", CellIsRule(operator="equal", formula=['"No"'], fill=PatternFill("solid", fgColor="F3D6CE")))
    # count block
    rr = last+2
    ws.cell(row=rr, column=1, value="Implementation snapshot").font = Font(name=FONT, bold=True, size=12, color=DEEP)
    for k,(lab,val) in enumerate([("Implemented (Yes)",'"Yes"'),("Partly",'"Partly"'),("Not yet (No)",'"No"')]):
        ws.cell(row=rr+1+k, column=1, value=lab); body(ws.cell(row=rr+1+k,column=1))
        cell = ws.cell(row=rr+1+k, column=2, value=f'=COUNTIF(E5:E{last},{val})'); body(cell, align="center", bold=True)
    note = ws.cell(row=rr+5, column=1,
        value="This is a representative starter set, not all 287 Moderate-baseline controls. It shows the shape of a control tracker: "
              "start from the baseline, tailor to your system, then track status and ownership. Full baselines are in SP 800-53B / OSCAL.")
    note.font = Font(name=FONT, italic=True, size=10, color=MUTED); note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=rr+5, start_column=1, end_row=rr+5, end_column=7)
    widths(ws, {"A":11,"B":34,"C":34,"D":11,"E":12,"F":16,"G":44})
    ws.freeze_panes = "A5"
    name = "control-selection-BLANK.xlsx" if blank else "control-selection.xlsx"
    wb.save(outpath("episode-03-categorize-select", name)); print(name)

# ============================================================ RMF step tracker
RMF = [
 ("1","Prepare","Get ready: know the system, its people, and its risk context.","Done","CISO","Portal team and data owner identified"),
 ("2","Categorize","Rate the system Low, Moderate, or High (FIPS 199).","Done","Head of Risk","Result: Moderate (see FIPS 199 workbook)"),
 ("3","Select","Choose the control baseline and tailor it.","In progress","Security team","Moderate baseline, tailoring underway"),
 ("4","Implement","Put the controls in place and write down how.","In progress","Platform team","Most technical controls exist, documenting"),
 ("5","Assess","Test that the controls actually work.","Not started","Security team","Assessment planned after implement"),
 ("6","Authorize","A responsible leader accepts the risk and signs off.","Not started","CISO","Awaiting assessment results"),
 ("7","Monitor","Keep watching, and re-check as things change.","Not started","Security team","Continuous monitoring being defined"),
]
def rmf_tracker():
    wb = Workbook(); ws = wb.active; ws.title = "RMF Tracker"
    title_bar(ws, "Meridian Pay  -  RMF Progress: Customer Web Portal (S2)", 6,
              "The seven steps of the Risk Management Framework (SP 800-37r2), tracked for one system.")
    header_row(ws, 4, ["Step","Name","What it means","Status","Owner","Notes"])
    r = 5
    for num, name, mean, status, owner, note in RMF:
        for c,v in enumerate([num,name,mean,status,owner,note],1):
            cell = ws.cell(row=r, column=c, value=v); body(cell, align="center" if c in (1,4) else "left")
            if r%2==0: cell.fill = PatternFill("solid", fgColor=ROW)
        ws.cell(row=r, column=2).font = Font(name=FONT, bold=True, size=10, color=DEEP)
        r += 1
    last = r-1
    dv = DataValidation(type="list", formula1='"Not started,In progress,Done"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"D5:D{last}")
    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"Done"'], fill=PatternFill("solid", fgColor="D8E6DC")))
    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"In progress"'], fill=PatternFill("solid", fgColor="F6E6C8")))
    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"Not started"'], fill=PatternFill("solid", fgColor="ECEFEA")))
    widths(ws, {"A":7,"B":16,"C":44,"D":14,"E":16,"F":40})
    ws.freeze_panes = "A5"
    wb.save(outpath("episode-02-rmf", "rmf-step-tracker.xlsx")); print("rmf-step-tracker.xlsx")

# ============================================================ POA&M
POAM = [
 ("P-001","Audit logs not reviewed on a schedule (AU-6)","CSF self-assessment","Moderate","Admin Console (S7)","Define weekly log-review process and assign it","Security team","Open"),
 ("P-002","Backup restores not tested regularly (CP-10)","CSF self-assessment","Moderate","Core Banking (S3)","Add quarterly restore test to the calendar","Platform team","Open"),
 ("P-003","No formal patch SLA (SI-2)","Risk assessment","Moderate","All systems","Set and document patch timelines by severity","Platform team","In progress"),
 ("P-004","Quarterly access review not yet formal (AC-2)","CSF self-assessment","Low","All systems","Stand up quarterly access review with owners","Platform team","Open"),
 ("P-005","Supplier risk assessments incomplete (GV.SC)","CSF self-assessment","High","Third parties","Assess CardaX, VerifyNow, messaging provider","Head of Risk","Open"),
]
def poam():
    wb = Workbook(); ws = wb.active; ws.title = "POA&M"
    title_bar(ws, "Meridian Pay  -  Plan of Action and Milestones (POA&M)", 9,
              "The master list of known weaknesses and how they get fixed. This is the working heart of a security programme.")
    header_row(ws, 4, ["ID","Weakness","Source","Severity","Affected system","Planned remediation","Owner","Target date","Status"])
    r = 5
    for pid, weak, src, sev, sysx, rem, owner, status in POAM:
        vals = [pid, weak, src, sev, sysx, rem, owner, "", status]
        for c,v in enumerate(vals,1):
            cell = ws.cell(row=r, column=c, value=v); body(cell, align="center" if c in (1,4,8,9) else "left")
            if r%2==0: cell.fill = PatternFill("solid", fgColor=ROW)
        r += 1
    # add blank rows for practice
    for _ in range(8):
        for c in range(1,10):
            cell = ws.cell(row=r, column=c, value=""); body(cell, align="center" if c in (1,4,8,9) else "left")
        r += 1
    last = r-1
    dvs = DataValidation(type="list", formula1='"High,Moderate,Low"', allow_blank=True); ws.add_data_validation(dvs); dvs.add(f"D5:D{last}")
    dvst = DataValidation(type="list", formula1='"Open,In progress,Done,Accepted"', allow_blank=True); ws.add_data_validation(dvst); dvst.add(f"I5:I{last}")
    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="equal", formula=['"High"'], fill=PatternFill("solid", fgColor="F3D6CE")))
    ws.conditional_formatting.add(f"I5:I{last}", CellIsRule(operator="equal", formula=['"Done"'], fill=PatternFill("solid", fgColor="D8E6DC")))
    widths(ws, {"A":8,"B":40,"C":20,"D":11,"E":18,"F":40,"G":16,"H":13,"I":13})
    ws.freeze_panes = "A5"
    wb.save(outpath("templates", "poam-tracker.xlsx")); print("poam-tracker.xlsx")

# ============================================================ gap assessment
GAP = [
 ("Oversight (GV.OV)","Leadership does not yet review security metrics regularly","Quarterly security review with the CTO","2","High","Medium"),
 ("Detection (DE.CM/DE.AE)","Monitoring exists but alerts are not analysed consistently","A defined monitoring and triage process","2","High","High"),
 ("Supply chain (GV.SC)","Providers not formally assessed","Documented assessments for key providers","2","High","Medium"),
 ("Improvement (ID.IM)","Lessons from incidents not fed back systematically","A blameless review after every incident, tracked","2","Medium","Low"),
 ("Access review (PR.AA)","Access granted well, but not reviewed on a schedule","Quarterly access recertification","1","Medium","Low"),
 ("Recovery testing (RC.RP)","Backups run but restores are rarely tested","Scheduled restore tests with evidence","2","High","Medium"),
]
def gap():
    wb = Workbook(); ws = wb.active; ws.title = "Gap Assessment"
    title_bar(ws, "Meridian Pay  -  Gap Assessment", 6,
              "Where we are, where we want to be, and how big the gap is. Sort by gap and priority to plan the roadmap.")
    header_row(ws, 4, ["Area","Current state","Desired state","Gap (1-4)","Priority","Effort"])
    r = 5
    for area, cur, des, g, prio, eff in GAP:
        for c,v in enumerate([area,cur,des,g,prio,eff],1):
            cell = ws.cell(row=r, column=c, value=v); body(cell, align="center" if c in (4,5,6) else "left")
            if r%2==0: cell.fill = PatternFill("solid", fgColor=ROW)
        r += 1
    for _ in range(6):
        for c in range(1,7):
            cell = ws.cell(row=r, column=c, value=""); body(cell, align="center" if c in (4,5,6) else "left")
        r += 1
    last = r-1
    dvp = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True); ws.add_data_validation(dvp); dvp.add(f"E5:E{last}")
    dve = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True); ws.add_data_validation(dve); dve.add(f"F5:F{last}")
    ws.conditional_formatting.add(f"D5:D{last}", CellIsRule(operator="greaterThanOrEqual", formula=["2"], fill=PatternFill("solid", fgColor="F3D6CE")))
    widths(ws, {"A":26,"B":40,"C":40,"D":10,"E":11,"F":10})
    ws.freeze_panes = "A5"
    wb.save(outpath("templates", "gap-assessment.xlsx")); print("gap-assessment.xlsx")

if __name__ == "__main__":
    csf_profile(False); csf_profile(True)
    fips199(False); fips199(True)
    control_selection(False); control_selection(True)
    rmf_tracker()
    poam()
    gap()
    print("done")
